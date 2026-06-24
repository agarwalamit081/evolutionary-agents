"""Tool extracting text from non-plain-text documents (PDF/DOCX/XLSX/CSV/HTML).

A dedicated parser is faster, safer, and cheaper than wrapping ``pypdf``/
``python-docx``/``openpyxl`` in a ``code_executor`` script every time. Sandbox
path guard reused from the ``file_reader`` pattern.

HTML (D1) reuses the ``web_scraper`` trafilatura→markdownify main-content chain.
The opt-in PDF rich path (D3) uses ``pymupdf`` (``fitz``) to extract tables and
render image-bearing pages to PNG; the default PDF path stays ``pypdf`` text-only.
"""

from __future__ import annotations

import asyncio
import csv
from pathlib import Path
from typing import Any, Optional, cast

from loguru import logger

from src.config.settings import get_settings


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _extract_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"### Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [("" if c is None else str(c)) for c in row]
            if any(cell.strip() for cell in cells):
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def _extract_csv(path: Path) -> str:
    with path.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        return "\n".join(",".join(row) for row in reader)


def _extract_html(target: Path) -> str:
    """HTML → markdown via the shared ``web_scraper`` trafilatura→markdownify chain.

    Reuses ``web_scraper._extract_markdown`` so the parser and the scraper agree on
    one main-content extraction path (D1). trafilatura is the primary extractor;
    ``markdownify`` is its optional fallback. Returns ``""`` when nothing could be
    extracted (``_extract`` then surfaces the no-text ``ERROR:``).
    """
    from src.tools.builtin.web_scraper import _extract_markdown

    html = target.read_text(encoding="utf-8", errors="replace")
    return _extract_markdown(html)


def _rows_to_markdown(rows: list[list[object]]) -> str:
    """Render fitz table rows (rows of cells) as a GitHub-flavored markdown table.

    ``None`` cells become empty; embedded newlines are flattened and literal
    ``|`` escaped so a cell never breaks the table grid.
    """
    if not rows:
        return ""

    def cell(value: object) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", " ").replace("|", "\\|")

    width = max(len(r) for r in rows)
    lines = ["| " + " | ".join(cell(c) for c in rows[0]) + " |"]
    lines.append("| " + " | ".join("---" for _ in range(width)) + " |")
    for row in rows[1:]:
        lines.append("| " + " | ".join(cell(c) for c in row) + " |")
    return "\n".join(lines)


def _extract_pdf_rich(target: Path, figures_dir: Path) -> str:
    """Full PDF extraction via ``pymupdf`` (``fitz``): text + tables + figure pages.

    Richer than the pypdf text-only path (``_extract_pdf``): extracts tables as
    markdown and renders image-bearing pages to PNG under ``figures_dir``. This is
    the opt-in path (``extract_figures=True``); the default stays pypdf text-only.

    If ``fitz`` is unavailable, degrades to the pypdf text path so the tool never
    hard-fails on a missing optional extraction lib.
    """
    try:
        import fitz  # pymupdf — installed + allowlisted (requirements.txt)
    except ImportError:
        return _extract_pdf(target)

    # pymupdf ships incomplete stubs (``Document`` is not declared Iterable),
    # so cast to Any — iteration/indexing are valid at runtime.
    doc = cast(Any, fitz.open(str(target)))
    parts: list[str] = []
    saved: list[str] = []
    try:
        for idx, page in enumerate(doc, start=1):
            text = (page.get_text("text") or "").strip()
            if text:
                parts.append(f"### Page {idx}\n{text}")
            # Tables on this page (best-effort: never abort parsing on a table error).
            try:
                finder = page.find_tables()
                tables = list(getattr(finder, "tables", []) or [])
            except Exception:
                tables = []
            for t_idx, tbl in enumerate(tables, start=1):
                rendered = _rows_to_markdown(tbl.extract())
                if rendered.strip():
                    parts.append(f"#### Table {t_idx} (page {idx})\n{rendered}")
            # Figures: render pages that embed images to PNG.
            if page.get_images():
                png = figures_dir / f"page_{idx}.png"
                try:
                    # Created lazily: a text/table-only PDF leaves no empty dir.
                    figures_dir.mkdir(parents=True, exist_ok=True)
                    pix = page.get_pixmap(dpi=150)
                    pix.save(str(png))
                    saved.append(str(png))
                except Exception as exc:  # never abort parsing on a render failure
                    logger.debug(f"document_parser: render page {idx} failed: {exc}")
    finally:
        doc.close()
    if saved:
        parts.append("#### Extracted figures (PNG)\n" + "\n".join(f"- {p}" for p in saved))
    return "\n\n".join(p for p in parts if p.strip())


def _extract(
    target: Path, ext: str, max_chars: int, extract_figures: bool = False, figures_dir: Path | None = None
) -> str:
    """Dispatch by extension (sync; run off the event loop)."""
    try:
        if ext == ".pdf":
            if extract_figures and figures_dir is not None:
                text = _extract_pdf_rich(target, figures_dir)
            else:
                text = _extract_pdf(target)
        elif ext == ".docx":
            text = _extract_docx(target)
        elif ext == ".xlsx":
            text = _extract_xlsx(target)
        elif ext == ".csv":
            text = _extract_csv(target)
        elif ext in (".html", ".htm"):
            text = _extract_html(target)
        elif ext in (".txt", ".md", ".text", ".markdown"):
            text = target.read_text(encoding="utf-8", errors="replace")
        else:
            return f"ERROR: Unsupported file type '{ext}'. Supported: pdf, docx, xlsx, csv, html, txt, md."
    except Exception as exc:
        return f"ERROR: Failed to parse {target.name}: {exc}"

    text = text.strip()
    if not text:
        return f"ERROR: No extractable text in {target.name}"
    if len(text) > max_chars:
        text = text[:max_chars] + f"\n\n... (truncated at {max_chars} chars)"
    return text


async def document_parser(
    file_path: str,
    max_chars: int = 8000,
    sandbox_root: Optional[str] = None,
    extract_figures: bool = False,
) -> str:
    """Extract text from a PDF, Word, Excel, CSV, HTML, or plain-text document.

    Args:
        file_path: Relative path to the document within the sandbox root.
        max_chars: Maximum characters of extracted text to return (default 8000).
        sandbox_root: Root directory for sandboxing (prevents path traversal).
            Defaults to ``settings.agent.workspace_root``.
        extract_figures: PDF only (default ``False``). When ``True``, extract tables
            as markdown and render image-bearing pages to PNG via pymupdf (``fitz``).
            The default PDF path stays text-only (``pypdf``).

    Returns:
        Extracted text, or an ``ERROR:`` string. Path traversal is blocked.
    """
    if sandbox_root is None:
        sandbox_root = get_settings().agent.workspace_root
    root = Path(sandbox_root).resolve()
    target = (root / file_path).resolve()

    if not target.is_relative_to(root):
        return f"ERROR: Path traversal blocked: {file_path}"
    if not target.exists():
        return f"ERROR: File not found: {file_path}"
    if not target.is_file():
        return f"ERROR: Not a file: {file_path}"

    # Cap input size at 20 MB to bound parsing cost.
    if target.stat().st_size > 20_000_000:
        return f"ERROR: File too large ({target.stat().st_size} bytes, max 20 MB)"

    figures_dir: Path | None = None
    if extract_figures and target.suffix.lower() == ".pdf":
        # Lazy import keeps the builtin importable when _paths is absent; resolves
        # to the per-run subdir when RESULTS_PER_RUN_SUBDIR is on, else results root.
        from src.tools._paths import results_root

        figures_dir = results_root() / "figures"

    logger.info(f"document_parser: {file_path}")
    ext = target.suffix.lower()
    return await asyncio.to_thread(_extract, target, ext, max_chars, extract_figures, figures_dir)


TOOL_DEFINITION = {
    "name": "document_parser",
    "handler": document_parser,
    "description": (
        "Extract text from a document: PDF, Word (.docx), Excel (.xlsx), CSV, "
        "HTML (.html/.htm), or plain text/markdown. HTML is converted to markdown "
        "via the main-content extractor. Faster and safer than writing a parsing "
        "script via code_executor. Path traversal is blocked. Set extract_figures "
        "to true to also extract PDF tables as markdown and render image-bearing "
        "pages to PNG (pymupdf); the default PDF path is text-only (pypdf)."
    ),
    # Deterministic per (path, max_chars, extract_figures) — safe to cache.
    "cacheable": True,
    "parameters": {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "Relative path to the document within the project directory.",
            },
            "max_chars": {
                "type": "integer",
                "description": "Maximum characters of extracted text (default: 8000).",
                "default": 8000,
            },
            "extract_figures": {
                "type": "boolean",
                "description": (
                    "PDF only. Extract tables as markdown and render image-bearing "
                    "pages to PNG under the results figures directory (uses pymupdf). "
                    "Default false (text-only via pypdf)."
                ),
                "default": False,
            },
        },
        "required": ["file_path"],
    },
}
